import json
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from collections import OrderedDict
import dateutil.parser


def process():
    with open("blockchain.json") as f:
        data = json.load(f, object_pairs_hook=OrderedDict)
    new_data = dict()
    for entry in data:
        if entry["repo"] not in new_data:
            new_data[entry["repo"]] = dict()
            new_data[entry["repo"]]["commits"] = 1
            new_data[entry["repo"]]["author"] = set()
            new_data[entry["repo"]]["author"].add(entry["email"])
        else:
            new_data[entry["repo"]]["commits"] += 1
            new_data[entry["repo"]]["author"].add(entry["email"])
    return new_data


def set_default(obj):
    if isinstance(obj, set):
        return list(obj)
    raise TypeError


def avg(commits, first, last):
    return commits / (abs((dateutil.parser.parse(last) - dateutil.parser.parse(first)).days) + 1)


# with open("c_a.json") as f:
#     new_data = json.load(f)
new_data = process()
c_a_data = dict()
for entry in new_data:
    contributor = len(new_data[entry]["author"])
    c_a_data[entry] = [contributor, new_data[entry]["commits"]]
with open("c_a_data.json", 'w+') as f:
    json.dump(c_a_data,f)

with open("c_a_data.json") as f:
    c_a_data = json.load(f)

with open('bitcoin.json') as f:
    bc_dir2project = json.load(f)

with open('ethereum.json') as f:
    eth_dir2project = json.load(f)

with open("repos/bitcoin.json") as f:
    data = json.load(f, object_pairs_hook=OrderedDict)

btc_add = list()
with open('add_btc.md') as f:
    for line in f.readlines():
        btc_add.append(line.strip('\n'))

eth_add = list()
with open('add_eth.md') as f:
    for line in f.readlines():
        eth_add.append(line.strip('\n'))
black_list = {'Tex', 'Jupyter Notebook', 'null'}
# print(json.dumps(data, indent=2))
# stargazers_count, open_issues_count, forks, pushed_at, watchers_count, updated_at, created_at, watchers, forks_count, open_issues, size
wb = Workbook()
dest_filename = 'blockchain.xlsx'
ws = wb.create_sheet(index=0, title='bitcoin')
attributes = ["watchers_count", "forks_count", "size", "pushed_at", "updated_at", "created_at"]
row = ["name"] + attributes + ["frequency", "contributors", "commits"]
ws.append(row)
eth_list = list()
btc_list = list()
for entry in data:
    if data[entry]["watchers_count"] < 5 and data[entry]["forks_count"] < 5 and entry not in btc_add:
        continue
    if data[entry]["language"] not in black_list and data[entry]["language"] is not None:
        btc_list.append(entry)
        row = [entry]
        for attr in attributes:
            row.append(data[entry][attr])
        for item in bc_dir2project:
            if item["project_name"] == entry:
                try:
                    row.append(avg(c_a_data[item["dir_name"]][1], data[entry]["created_at"], max(data[entry]["pushed_at"], data[entry]["updated_at"])))
                    row.append(c_a_data[item["dir_name"]][0])
                    row.append(c_a_data[item["dir_name"]][1])
                except BaseException:
                    pass
                break
        ws.append(row)
with open("repos/ethereum.json") as f:
    data = json.load(f, object_pairs_hook=OrderedDict)
ws = wb.create_sheet(index=0, title='ethereum')
row = ["name"] + attributes + ["frequency", "contributors", "commits"]
ws.append(row)

for entry in data:
    if data[entry]["watchers_count"] < 5 and data[entry]["forks_count"] < 5 and entry not in eth_add:
        continue
    if data[entry]["language"] not in black_list and data[entry]["language"] is not None:
        eth_list.append(entry)
        row = [entry]
        for attr in attributes:
            row.append(data[entry][attr])
        for item in eth_dir2project:
            if item["project_name"] == entry:
                try:
                    row.append(avg(c_a_data[item["dir_name"]][1], data[entry]["created_at"],
                                   max(data[entry]["pushed_at"], data[entry]["updated_at"])))
                    row.append(c_a_data[item["dir_name"]][0])
                    row.append(c_a_data[item["dir_name"]][1])
                except BaseException:
                    pass
                break
        ws.append(row)
wb.save(dest_filename)

with open('eth_selected.json', 'w+') as f:
    json.dump(eth_list, f)

with open('btc_selected.json', 'w+') as f:
    json.dump(btc_list, f)
# def sort_by_attribute(data, attribute):
#     return sorted(data.items(), key=lambda x: x[1][attribute], reverse=True)
#
#
# print(json.dumps(sort_by_attribute(data, "stargazers_count"), indent=2))